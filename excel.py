#!/usr/bin/python3
# excel.py
# Author: Claudio <3261958605@qq.com>
# Created: 2017-07-26 09:44:34
# Code:
'''
excel表格操作
'''

# 针对本项目的工具函数

import datetime
import itertools
import openpyxl as xls
import public
import re
from collections import namedtuple
from openpyxl.styles import Alignment, Font, NamedStyle

#
# API函数
#


# 表格中的头信息
HEADER = {
    'title': '收费站鲜活车登记表',
    'company': '单位：XXX收费站',
    'headers': ['序号', '日期', '时间', '车道', '车牌', '车型',
                '农产品种类', '操作人', '照片索引', '备注']
}

# 表格中的每行内容
Entry = namedtuple(
    "Entry",
    'index date time lane plate car_type product operator pics_index notes')
# index：索引（计算获取）
# date：图片创建日期，格式为2017.07.01
# time：图片创建创建时间（精确到分），格式为17:09(24小时制)
# lane：车道编号
# plate：车牌（不需）
# car_type：车型 1,2,3...（默认）
# product：农产品种类（不需）
# operator：操作人（不需）
# pics_index：照片索引（添加行时自动创建）：格式为20170701001.1-001.3
# notes：备注（不需）

# 日期和时间栏的文本匹配正则
# 日期格式：2017.07.01
# 时间格式：03:21:00 AM
DATE_REGEX = re.compile(r'(?P<year>\d{4}).(?P<month>\d{2}).(?P<day>\d{2})')

# 时间为datetime.time对象，极少可能因为偷懒使用了字符串
TIME_REGEX = re.compile(
    r'(?P<hour>\d{2})[:;；](?P<minute>\d{2})',
    re.IGNORECASE)

# 日期和时间所在行
COL_DATE = 2
COL_TIME = 3

CELL_STYLE = NamedStyle(name="cell_style")
CELL_STYLE.alignment = Alignment(horizontal='center')
CELL_STYLE.font = Font(size=12, name='DejaVu Sans')


def _load_book(filename):
    '仅做测试用'
    import os.path
    if os.path.exists(filename):
        return xls.load_workbook(filename)
    else:
        book = xls.Workbook()
        book.save(filename)
        return book


def get_sheet(book, month):
    name = '{}月'.format(month)
    # print('表单名称：', name)

    if name in book.get_sheet_names():
        return book.get_sheet_by_name(name)
    else:
        return create_sheet(book)


def get_sheets(book):
    '为跨月操作，需获取2个表单'
    sheets = []
    date_current = public.date_current()
    month = date_current.month

    # 前一个月的表单
    if 1 == date_current.day and 1 != month:
        sheets.append(get_sheet(book, month - 1))
    else:
        sheets.append(None)

    # 当月表单
    sheets.append(get_sheet(book, month))

    return sheets


def create_sheet(book):
    name = '{}月'.format(public.now().month)

    sheet = book.create_sheet(name)

    # 添加头信息

    for row in [1, 2]:
        sheet.merge_cells(
            start_row=row, end_row=row,
            start_column=1, end_column=len(HEADER['headers']) - 1)

    sheet.cell(row=1, column=1, value=HEADER['title'])
    sheet.cell(row=2, column=1, value=HEADER['company'])

    headers = HEADER['headers']
    for i, h in enumerate(headers):
        sheet.cell(row=3, column=i + 1, value=headers[i])

    return sheet


def get_datetime(row):
    '''
    获取行中的日期
    如果能成功获取，返回datetime对象，
    否则返回None
    '''
    date_value = ''
    time_value = ''
    is_str = False              # 时间是否为字符串类型
    for col, cell in enumerate(row):
        if not cell.value:
            break

        col += 1
        if col == COL_DATE:
            date_value = cell.value
        if col == COL_TIME:
            time_value = cell.value
            break

    if not (date_value and time_value):
        return None

    #
    # 判断时间类型，检查是否为空字符串
    #

    # 日期 str类型
    date_value = date_value.strip()
    # 时间 可能是str类型，也可能是datetime.time类型
    if isinstance(time_value, str):
        time_value = time_value.strip()
        is_str = True
        if (not date_value) or (not time_value):
            return None

    #
    # 确定时间和日期
    #
    mo_date = DATE_REGEX.match(date_value)
    if not mo_date:
        return None
    year = int(mo_date.group('year'))
    month = int(mo_date.group('month'))
    day = int(mo_date.group('day'))

    if is_str:
        # for cell in row:
        #     print(cell.value, end=' ')
        # print()
        # 使用search，而不是match，为了放置手动输入了秒
        mo_time = TIME_REGEX.search(time_value)
        if not mo_time:
            return None
        hour = int(mo_time.group('hour'))
        minute = int(mo_time.group('minute'))
        second = 0
    elif isinstance(time_value, datetime.time):
        hour = time_value.hour
        minute = time_value.minute
        second = time_value.second
    else:
        return None

    return datetime.datetime(year=year, month=month,
                             day=day, hour=hour,
                             minute=minute, second=second)


def get_edited_and_row_start(excel_file, guard_datetime):
    '''
    获取已编辑车辆数，和开始插入的行
    @excel_file：主要是sheet主要为了为operation.py提供接口
    @not_edited：昨天没处理的车辆数
    返回值：元组(昨天处理过的车辆数， 开始写入的行)
    '''
    book = xls.load_workbook(excel_file)

    # 确定是否操作前一个的表单
    sheets = get_sheets(book)
    sheet = sheets[0] or sheets[1]
    # 昨天出处理过的车辆数
    edited = 0

    # 开始写入的行
    row_start = 0

    # 有昨天条目的第一行
    row_of_yesterday_start = 0

    # 只需判断天数，所以重建
    date_current = public.date_current()

    #
    # 循环获已编辑昨天的行数，以及昨天条目的第一行
    #

    for r, row in enumerate(sheet.rows):
        row_datetime = get_datetime(row)
        if not row_datetime:
            continue

        row_date = datetime.datetime(
            year=row_datetime.year,
            month=row_datetime.month,
            day=row_datetime.day)

        date_delta = date_current - row_date
        if 1 == date_delta.days:
            if 0 == row_of_yesterday_start:
                row_of_yesterday_start = r + 1

            if row_datetime < guard_datetime:
                edited += 1
                # print(r, row_datetime)

    # 确定开始写入的行
    # 如果没有昨天的行，有2种情况：
    # 1. 新建的的sheet
    # 2. 确实没有昨天的条目
    # print('row_of_yesterday_start：', row_of_yesterday_start, '最后行号：', r + 2)
    if 0 == row_of_yesterday_start:
        if len(HEADER) == sheet.max_row:
            row_start = len(HEADER) + 1
        else:
            row_start = r + 2
    else:
        row_start = row_of_yesterday_start + edited

    return (edited, row_start)


def add_entry(sheet, datetime_, row, default_lane='7', pics_index=''):
    '''向sheet中添加一行数据
    @datetime_：datetime.datetime对象
    @row：需插入的行
    @default_lane：字符串，默认车道

    date格式：2017.01.10
    time格式：为日期对象


    '''

    index = row - len(HEADER)
    date = '{}.{:02}.{:02}'.format(
        datetime_.year, datetime_.month, datetime_.day)
    time_ = datetime.time(hour=datetime_.hour,
                          minute=datetime_.minute, second=datetime_.second)
    # time_ = '{}:{}'.format(datetime_.hour, datetime_.minute)
    # second=datetime_.second)
    lane = default_lane
    plate = '川'
    car_type = 1
    product = ''
    operator = ''
    pics_index = pics_index
    notes = ''

    entry = Entry(index=index, date=date, time=time_, lane=lane,
                  plate=plate, car_type=car_type,
                  product=product, operator=operator,
                  pics_index=pics_index, notes=notes)

    for i, x in enumerate(entry):
        col = i + 1
        _cell = sheet.cell(column=col, row=row)
        _cell.value = x
        # 要改变Excel中显示格式还真不容易
        # _cell.number_format = 'xxx'还不能写在_cell.value=x前面
        # 参考地址：https://stackoverflow.com/questions/24370385/how-to-format-cell-with-datetime-object-of-the-form-yyyy-mm-dd-hhmmss-in-exc
        if col == COL_TIME:
            _cell.number_format = 'HH:MM'

    return sheet


def add_entries(sheet, car_datetimes, row_start, default_lane='7', edited=0):
    '''

    添加完成后，删除后面所有行

    @car_datetimes:已经排序好的datetime对象

    pics_index：格式为“20170701001.1-001.3”
    '''

    date_current = public.date_current()
    index_yesterday = edited
    index_today = 0
    index_fake = 999
    index = 0

    for datetime_ in car_datetimes:
        date_car = datetime.datetime(
            datetime_.year, datetime_.month, datetime_.day)
        days = (date_current - date_car).days

        if 1 == days:
            index_yesterday += 1
            index = index_yesterday
        elif 0 == days:
            index_today += 1
            index = index_today
        else:
            index_fake += 1
            index = index_fake

        pics_index = '{}{:02}{:02}{:03}.1-{:03}.3'.format(
            datetime_.year, datetime_.month, datetime_.day,
            index, index)

        # print('第{}行'.format(row_start), end='\t')
        add_entry(sheet, datetime_, row_start,
                  default_lane=default_lane, pics_index=pics_index)

        row_start += 1

    # 在重复修改时，如果有删减车辆数，可去除后面多余行
    clear_rows_from(sheet, row_start)


def add_entries_for_sheets(sheets, car_datetimes, row_start, default_lane='7',
                           edited=0, not_edited=0):
    '''为当月和前一月表单中添加数据
    参数与add_entries相同不过row_start为第一个表单中开始插入的行
    '''
    sheet1 = sheets[0]            # 前一月的表单
    sheet2 = sheets[1]            # 当月表单

    # 如需跨月操作
    if sheet1:
        car_datetimes1 = car_datetimes[0:not_edited]
        car_datetimes2 = car_datetimes[not_edited:]
        add_entries(sheet1, car_datetimes1, row_start,
                    default_lane='7', edited=edited)
        add_entries(sheet2, car_datetimes2, len(HEADER) + 1,
                    default_lane='7', edited=0)
    # 否则
    else:
        add_entries(sheet2, car_datetimes, row_start,
                    default_lane='7', edited=edited)


def clear_row(sheet, row):
    col_start = 1
    for i in range(len(HEADER['headers'])):
        _ = sheet.cell(column=col_start + i, row=row, value='')


def clear_rows_from(sheet, row):
    for r in itertools.count(row):
        first_cell = sheet.cell(row=r, column=1)
        if not first_cell.value:
            break
        else:
            clear_row(sheet, r)


#
# 此模块总函数
#

def main(excel_file, car_datetimes, row_start, default_lane='7',
         edited=0, not_edited=0):
    '''
    @excel_file：需操作的Excel文件
    @car_datetimes：datetime对象组成的序列
    @row_start：开始插入的行
    @default_lane：车道编号
    @edited：昨天编辑好的车辆数
    @not_edited：昨天没处理的车辆数
    '''
    try:
        book = xls.load_workbook(excel_file)
        sheets = get_sheets(book)
        add_entries_for_sheets(sheets,
                               car_datetimes,
                               row_start,
                               default_lane=default_lane,
                               edited=edited,
                               not_edited=not_edited)
        book.save(excel_file)
    except:
        return '{}：操作Excel文件失败，请确保没有其他程序占用该文件'.format(excel_file)

    return True
